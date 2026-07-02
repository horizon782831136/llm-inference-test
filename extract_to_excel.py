#!/usr/bin/env python3
"""Extract Benchmarking summary tables from log_rate_perf.log into an Excel file.

Each summary table becomes one row in the output spreadsheet. Columns cover
General / Latency / Tokens / Speculative Decoding metrics, plus the
``Running with rate=..., parallel=...`` context that precedes each block.
"""
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

LOG_PATH = Path("/Users/liuwei/Desktop/Test/0630/log_8830_2.log")
OUT_PATH = Path("/Users/liuwei/Desktop/Test/0630/log_perf_8830_2.xlsx")

ANSI_RE = re.compile(r"\x1b\[[0-9;]*m")

# Metric label -> column header (kept in display order)
METRIC_COLUMNS = [
    ("Test Duration (s)", "Test Duration (s)"),
    ("Concurrency", "Concurrency"),
    ("Request Rate (req/s)", "Request Rate (req/s)"),
    ("Total / Success / Failed", "Total / Success / Failed"),
    ("Req Throughput (req/s)", "Req Throughput (req/s)"),
    ("Avg Latency (s)", "Avg Latency (s)"),
    ("TTFT (ms)", "TTFT (ms)"),
    ("TPOT (ms)", "TPOT (ms)"),
    ("ITL (ms)", "ITL (ms)"),
    ("Avg Input Tokens", "Avg Input Tokens"),
    ("Avg Output Tokens", "Avg Output Tokens"),
    ("Output Throughput (tok/s)", "Output Throughput (tok/s)"),
    ("Total Throughput (tok/s)", "Total Throughput (tok/s)"),
    ("Decoded Tok/Iter", "Decoded Tok/Iter"),
    ("Spec. Accept Rate", "Spec. Accept Rate"),
]


def parse_value(raw: str):
    """Convert a metric cell value to int/float when possible, else keep str."""
    raw = raw.strip()
    if not raw:
        return ""
    # Composite values like "200 / 200 / 0" — keep as string.
    if "/" in raw:
        return raw
    try:
        if "." in raw:
            return float(raw)
        return int(raw)
    except ValueError:
        return raw


def extract_rows(log_text: str):
    lines = log_text.splitlines()
    rows = []
    current_run = {"rate": None, "parallel": None}
    run_re = re.compile(r"Running with rate=([\d.]+),\s*parallel=(\d+)")
    # A summary table starts with this header line.
    header_marker = "│ Metric"
    end_marker_prefix = "└"

    i = 0
    while i < len(lines):
        line = ANSI_RE.sub("", lines[i])

        m = run_re.search(line)
        if m:
            current_run = {"rate": float(m.group(1)), "parallel": int(m.group(2))}

        # Detect the Benchmarking summary table — header row contains "│ Metric"
        # and the immediately preceding context is "Benchmarking summary:".
        if header_marker in line and "Value" in line:
            # Scan back a few lines to confirm it's the benchmarking summary
            # (the percentile table also starts with "│ Metric").
            preceding = "\n".join(ANSI_RE.sub("", lines[j]) for j in range(max(0, i - 4), i))
            if "Benchmarking summary" not in preceding:
                i += 1
                continue

            row = {col: "" for _, col in METRIC_COLUMNS}
            row["Rate (req/s) [config]"] = current_run["rate"]
            row["Parallel [config]"] = current_run["parallel"]

            # Walk forward until the table's bottom border.
            j = i + 1
            while j < len(lines):
                inner = ANSI_RE.sub("", lines[j])
                if inner.lstrip().startswith(end_marker_prefix):
                    break
                # Skip separator / section header rows.
                if "├" in inner or "──" in inner.replace("│", "").replace(" ", "")[:6]:
                    # Section heading rows look like "│ ── General ── │ │"; skip.
                    j += 1
                    continue
                # Match a data row: │ <label> │ <value> │
                cells = [c.strip() for c in inner.strip().strip("│").split("│")]
                if len(cells) == 2:
                    label, value = cells
                    # Skip the section dividers like "── General ──"
                    if label.startswith("──"):
                        j += 1
                        continue
                    for metric_label, column in METRIC_COLUMNS:
                        if label == metric_label:
                            row[column] = parse_value(value)
                            break
                j += 1

            rows.append(row)
            i = j + 1
            continue

        i += 1

    return rows


def write_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark Summary"

    headers = ["#", "Rate (req/s) [config]", "Parallel [config]"] + [col for _, col in METRIC_COLUMNS]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    center = Alignment(horizontal="center", vertical="center")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for idx, row in enumerate(rows, start=1):
        ws.append([idx] + [row.get(h, "") for h in headers[1:]])

    # Auto-size columns based on the longest value in each.
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2

    ws.freeze_panes = "B2"
    wb.save(OUT_PATH)


def main():
    text = LOG_PATH.read_text(encoding="utf-8", errors="replace")
    rows = extract_rows(text)
    print(f"Extracted {len(rows)} benchmark summary rows")
    write_excel(rows)
    print(f"Saved -> {OUT_PATH}")


if __name__ == "__main__":
    main()
